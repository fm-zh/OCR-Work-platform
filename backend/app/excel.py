"""把辨識文字用 DeepSeek 整理成表格，再用 openpyxl 產生 .xlsx。"""
from __future__ import annotations

import io
import json
import re
import sys
import urllib.request
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import llm_correct  # noqa: E402
import openpyxl  # noqa: E402

_STRUCT_PROMPT = """以下是一頁台灣財務報表/稅報的辨識文字。請整理成一個表格的 JSON 物件，格式：
{{"columns": ["欄1","欄2",...], "rows": [["儲存格",...], ...]}}

規則：
- 欄位依內容判斷（例如「項目/金額/百分比」或「代碼/科目/金額」）；若難以判斷欄位可給空陣列 columns。
- 每一列對應原文一行的資料；把標籤與數字拆到對應欄位。
- 儲存格值「保留原始文字」：含千分位逗號、括號負號（如 (588)）、百分比 %、$ 等，不要改寫成數值。
- 只輸出單一 JSON 物件，不要任何說明或 markdown。

辨識文字：
=====
{text}
====="""


def _fallback(source_text: str) -> dict:
    rows = [[ln] for ln in source_text.splitlines() if ln.strip()]
    return {"columns": [], "rows": rows}


def _parse_structured(raw: str, source_text: str) -> dict:
    """DeepSeek 回傳字串 → {columns, rows}；失敗退化為每行單欄。"""
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return _fallback(source_text)
    if not isinstance(data, dict):
        return _fallback(source_text)
    rows = data.get("rows")
    if not isinstance(rows, list) or not all(isinstance(r, list) for r in rows):
        return _fallback(source_text)
    columns = data.get("columns")
    if not isinstance(columns, list):
        columns = []
    columns = [str(c) for c in columns]
    rows = [[("" if c is None else str(c)) for c in r] for r in rows]
    return {"columns": columns, "rows": rows}


def _deepseek_json(prompt: str, api_key: str, model: str, timeout: int) -> str:
    """呼叫 DeepSeek chat（response_format=json_object），回傳內容字串。"""
    body = json.dumps({
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "stream": False,
        "response_format": {"type": "json_object"},
    }).encode("utf-8")
    req = urllib.request.Request(
        f"{llm_correct.DEEPSEEK_API_BASE}/chat/completions",
        data=body, method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key.strip()}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        data = json.loads(r.read())
    return data["choices"][0]["message"]["content"]


def structure_page(text: str, api_key: str, model: str = "deepseek-chat",
                   timeout: int = 180) -> dict:
    """呼叫 DeepSeek 把一頁文字整理成 {columns, rows}。"""
    if not text or not text.strip():
        return {"columns": [], "rows": []}
    raw = _deepseek_json(_STRUCT_PROMPT.format(text=text), api_key, model, timeout)
    return _parse_structured(raw, text)


_LABEL_FIX_PROMPT = """以下是一頁台灣財務報表中「中文科目名稱」儲存格的清單（依表格由上而下排序），\
以 JSON 物件給出，key 為編號、value 為 OCR 文字。這些名稱可能被掃描裁切（左半遺失）\
或有形近/簡體誤字，例如：「當現金」應為「現金及約當現金」、「爭額」應為「應收帳款淨額」、\
「安公允價值衡量之金融資產流動」應為「透過損益按公允價值衡量之金融資產-流動」。

請逐一還原成完整、正確的台灣財報慣用科目名稱。規則：
- 只輸出單一 JSON 物件，key 與輸入「完全相同」，value 為修正後文字。
- 名稱後若帶附註參照（如「(附註五(一))」），保留該參照、只還原名稱本體。
- 只修明顯截斷或錯字；無把握就原樣回傳該 value。
- 不要新增 key、不要任何說明文字。

科目清單：
{cells}"""

# 「受保護」儲存格（保留原值、不交給 LLM 修改）：含阿拉伯數字（金額/年度/百分比），
# 或純中文數字附註代碼（如 四(十二)）。帶中文科目名的格（即使含括號附註）可被修正。
_HAS_DIGIT = re.compile(r"\d")
_NOTE_ONLY = re.compile(r"^[\s一二三四五六七八九十百千〇零()（）.、,$%:：－\-]+$")


def _is_protected(cell: str) -> bool:
    c = cell.strip()
    if not c:
        return True
    if _HAS_DIGIT.search(c):
        return True
    return bool(_NOTE_ONLY.match(c))


def restore_labels(grid: dict, api_key: str, model: str = "deepseek-chat",
                   timeout: int = 120) -> dict:
    """用 DeepSeek 修正表格中「截斷/錯字的中文科目名稱」，但不動數字與欄位結構。

    幾何重建已保證欄位正確；此步只補回被掃描裁切或誤判的中文標籤。任何會改變
    表格形狀、或更動到「含數字/符號/空白」儲存格的回應一律捨棄、退回原表，
    確保永不破壞已正確的數據。
    """
    rows = grid.get("rows") or []
    if not rows or not api_key or not api_key.strip():
        return grid
    # 只挑「中文科目名稱」格送修；含數字/符號/空白格完全不送、永不更動。
    cells = []  # (row_idx, col_idx, text)
    for ri, row in enumerate(rows):
        for ci, c in enumerate(row):
            c = "" if c is None else str(c)
            if c.strip() and not _is_protected(c):
                cells.append((ri, ci, c))
    if not cells:
        return grid
    payload = {str(i): t for i, (_, _, t) in enumerate(cells)}
    prompt = _LABEL_FIX_PROMPT.format(
        cells=json.dumps(payload, ensure_ascii=False))
    try:
        raw = _deepseek_json(prompt, api_key, model, timeout)
        fixes = json.loads(raw)
    except (RuntimeError, json.JSONDecodeError, TypeError, KeyError):
        return grid
    if not isinstance(fixes, dict):
        return grid
    # 回填到原表「固定位置」：結構/數字由原表決定，永不被破壞。
    out = [list(r) for r in rows]
    for i, (ri, ci, _orig) in enumerate(cells):
        nv = fixes.get(str(i))
        if isinstance(nv, str) and nv.strip():
            out[ri][ci] = nv
    return {"columns": grid.get("columns") or [], "rows": out}


_ILLEGAL_SHEET = re.compile(r"[\[\]:*?/\\]")


def _sanitize_sheet_title(name: str) -> str:
    """剔除 Excel 工作表名非法字元並截到 31 字；空字串以 'Sheet' 保底。"""
    s = _ILLEGAL_SHEET.sub("", name)[:31]
    return s or "Sheet"


def _sheet_name(page_no: int) -> str:
    return _sanitize_sheet_title(f"第{page_no}頁")


def merge_sheets(pages_in_order: list[dict]) -> dict:
    """把依頁碼排序的多頁 {columns, rows} 疊接清理成單一 {columns, rows}。

    規則（見設計 §合併演算法）：
    - 表頭取第一個 columns 非空的頁；全空則表頭為 []。
    - 欄數 W：表頭非空 → len(表頭)；表頭空 → 各頁所有列中最長列長度。
    - 逐頁疊接 rows：不附加任何頁的 columns；第一頁之後，若某頁第一列逐格
      去空白後等於合併表頭，視為重複表頭並丟棄；列短於 W 右補空，列長於 W 保留。
    """
    header: list[str] = []
    for p in pages_in_order:
        cols = [str(c) for c in (p.get("columns") or [])]
        if cols:
            header = cols
            break

    if header:
        width = len(header)
    else:
        width = 0
        for p in pages_in_order:
            for r in (p.get("rows") or []):
                width = max(width, len(r))

    header_trimmed = [h.strip() for h in header]
    out_rows: list[list[str]] = []
    for idx, p in enumerate(pages_in_order):
        rows = [[("" if c is None else str(c)) for c in r]
                for r in (p.get("rows") or [])]
        if idx > 0 and header and rows:
            if [c.strip() for c in rows[0]] == header_trimmed:
                rows = rows[1:]
        for r in rows:
            if len(r) < width:
                r = r + [""] * (width - len(r))
            out_rows.append(r)
    return {"columns": header, "rows": out_rows}


def build_workbook(pages_structured: dict, sheet_titles: dict | None = None) -> bytes:
    """{page_no: {columns, rows}} → .xlsx bytes（每頁一個工作表）。
    sheet_titles 給定時，對應 page_no 改用自訂工作表名（會自動剔除非法字元、截 31 字）。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for page_no in sorted(pages_structured):
        st = pages_structured[page_no]
        custom = (sheet_titles or {}).get(page_no)
        title = _sanitize_sheet_title(custom) if custom else _sheet_name(page_no)
        ws = wb.create_sheet(title=title)
        columns = st.get("columns") or []
        if columns:
            ws.append([str(c) for c in columns])
        for row in st.get("rows") or []:
            ws.append([("" if c is None else str(c)) for c in row])
    if not wb.sheetnames:
        wb.create_sheet(title="第1頁")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
