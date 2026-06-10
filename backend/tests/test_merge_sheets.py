from app.excel import merge_sheets


def test_basic_stacks_rows_and_keeps_first_header():
    pages = [
        {"columns": ["項目", "金額"], "rows": [["現金", "100"], ["應收", "200"]]},
        {"columns": ["項目", "金額"], "rows": [["存貨", "300"], ["合計", "600"]]},
    ]
    out = merge_sheets(pages)
    assert out["columns"] == ["項目", "金額"]
    assert out["rows"] == [["現金", "100"], ["應收", "200"],
                           ["存貨", "300"], ["合計", "600"]]


def test_drops_repeated_header_on_continuation_page():
    pages = [
        {"columns": ["項目", "金額"], "rows": [["現金", "100"]]},
        {"columns": ["項目", "金額"], "rows": [["項目", "金額"], ["存貨", "300"]]},
    ]
    out = merge_sheets(pages)
    assert out["rows"] == [["現金", "100"], ["存貨", "300"]]


def test_keeps_real_data_row_on_continuation_page():
    pages = [
        {"columns": ["項目", "金額"], "rows": [["現金", "100"]]},
        {"columns": ["項目", "金額"], "rows": [["存貨", "300"], ["合計", "600"]]},
    ]
    out = merge_sheets(pages)
    assert out["rows"][1] == ["存貨", "300"]  # 首列非表頭，未被刪


def test_pads_short_rows_and_keeps_long_rows():
    pages = [
        {"columns": ["a", "b", "c"], "rows": [["1"], ["1", "2", "3", "4"]]},
    ]
    out = merge_sheets(pages)
    assert out["rows"][0] == ["1", "", ""]            # 補空到 3
    assert out["rows"][1] == ["1", "2", "3", "4"]      # 長列保留不截


def test_header_taken_from_first_nonempty_columns():
    pages = [
        {"columns": [], "rows": [["x", "y"]]},
        {"columns": ["A", "B"], "rows": [["1", "2"]]},
    ]
    out = merge_sheets(pages)
    assert out["columns"] == ["A", "B"]


def test_all_columns_empty_uses_max_row_width():
    pages = [
        {"columns": [], "rows": [["1"]]},
        {"columns": [], "rows": [["1", "2", "3"]]},
    ]
    out = merge_sheets(pages)
    assert out["columns"] == []
    assert out["rows"][0] == ["1", "", ""]   # W = 最長列 3，短列補空


def test_single_page_and_empty():
    assert merge_sheets([]) == {"columns": [], "rows": []}
    one = merge_sheets([{"columns": ["a"], "rows": [["1"]]}])
    assert one == {"columns": ["a"], "rows": [["1"]]}


import io
import openpyxl
from app.excel import build_workbook


def test_build_workbook_custom_sheet_title():
    data = build_workbook({1: {"columns": ["a"], "rows": [["1"]]}},
                          sheet_titles={1: "2024財報"})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["2024財報"]


def test_build_workbook_custom_title_sanitized_and_truncated():
    long_name = "報表/名稱:" + "X" * 40
    data = build_workbook({1: {"columns": [], "rows": [["1"]]}},
                          sheet_titles={1: long_name})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    name = wb.sheetnames[0]
    assert "/" not in name and ":" not in name      # 非法字元剔除
    assert len(name) <= 31                            # Excel 上限


def test_build_workbook_default_naming_unchanged():
    data = build_workbook({3: {"columns": ["a"], "rows": [["1"]]}})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["第3頁"]
