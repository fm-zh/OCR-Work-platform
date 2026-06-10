import time
from pathlib import Path
from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)
BORN = r"E:\PJ_OCR\test\3.資產負債表測試-此為可複製文字之PDF.pdf"


def _upload():
    with open(BORN, "rb") as f:
        return client.post("/api/jobs",
                           files={"file": ("file3.pdf", f, "application/pdf")})


def test_health():
    assert client.get("/api/health").json() == {"status": "ok"}


def test_upload_creates_job():
    r = _upload()
    assert r.status_code == 200
    j = r.json()
    assert j["status"] == "created"
    assert j["n_pages"] == 1
    assert j["is_born_digital"] is True
    assert j["job_id"]


def test_reject_unsupported_extension():
    r = client.post("/api/jobs",
                    files={"file": ("bad.txt", b"hello", "text/plain")})
    assert r.status_code == 400


def test_page_image_returns_png():
    jid = _upload().json()["job_id"]
    r = client.get(f"/api/jobs/{jid}/pages/1/image")
    assert r.status_code == 200
    assert r.headers["content-type"] == "image/png"


def test_recognize_then_poll_done():
    jid = _upload().json()["job_id"]
    r = client.post(f"/api/jobs/{jid}/recognize", json={"pages": [1]})
    assert r.status_code == 200
    assert r.json()["status"] in ("running", "done")
    s = None
    for _ in range(60):
        s = client.get(f"/api/jobs/{jid}").json()
        if s["status"] in ("done", "error"):
            break
        time.sleep(0.3)
    assert s["status"] == "done"
    assert s["mode"] == "文字層擷取"
    assert s["pages"]["1"]


def test_unknown_job_404():
    assert client.get("/api/jobs/nope").status_code == 404


def test_delete_then_404():
    jid = _upload().json()["job_id"]
    assert client.delete(f"/api/jobs/{jid}").status_code == 204
    assert client.get(f"/api/jobs/{jid}").status_code == 404


import io
import openpyxl


def _excel_req(merge: bool):
    return {
        "file_name": "2024財報.pdf",
        "sheets": {
            "1": {"columns": ["項目", "金額"], "rows": [["現金", "100"]]},
            "2": {"columns": ["項目", "金額"], "rows": [["項目", "金額"], ["存貨", "300"]]},
        },
        "merge": merge,
    }


def test_excel_merge_true_single_sheet_named_by_file():
    r = client.post("/api/excel", json=_excel_req(True))
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["2024財報"]                 # 檔名主檔名命名
    ws = wb["2024財報"]
    # 表頭列 + 現金 + 存貨（第2頁重複表頭列被移除）
    values = [[c.value for c in row] for row in ws.iter_rows()]
    assert values == [["項目", "金額"], ["現金", "100"], ["存貨", "300"]]


def test_excel_merge_false_keeps_per_page_sheets():
    r = client.post("/api/excel", json=_excel_req(False))
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["第1頁", "第2頁"]


def test_excel_merge_defaults_false_when_omitted():
    body = _excel_req(False)
    del body["merge"]
    r = client.post("/api/excel", json=body)
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["第1頁", "第2頁"]
