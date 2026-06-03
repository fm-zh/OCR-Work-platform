import io
import openpyxl
from app import excel


def test_parse_structured_valid():
    raw = '{"columns":["項目","金額"],"rows":[["現金","100"]]}'
    out = excel._parse_structured(raw, "ignored")
    assert out["columns"] == ["項目", "金額"]
    assert out["rows"] == [["現金", "100"]]


def test_parse_structured_fallback_on_bad_json():
    out = excel._parse_structured("not json", "行一\n行二")
    assert out["columns"] == []
    assert out["rows"] == [["行一"], ["行二"]]


def test_parse_structured_coerces_non_strings():
    raw = '{"columns":["c"],"rows":[[1, null, "x"]]}'
    out = excel._parse_structured(raw, "")
    assert out["rows"] == [["1", "", "x"]]


def test_build_workbook_with_columns():
    data = {1: {"columns": ["項目", "金額"], "rows": [["現金", "100"], ["存貨", "200"]]}}
    wb = openpyxl.load_workbook(io.BytesIO(excel.build_workbook(data)))
    assert wb.sheetnames == ["第1頁"]
    ws = wb["第1頁"]
    assert [c.value for c in ws[1]] == ["項目", "金額"]
    assert ws.cell(row=2, column=1).value == "現金"


def test_build_workbook_multi_page_no_columns():
    data = {1: {"columns": [], "rows": [["a"]]}, 2: {"columns": [], "rows": [["b"]]}}
    wb = openpyxl.load_workbook(io.BytesIO(excel.build_workbook(data)))
    assert wb.sheetnames == ["第1頁", "第2頁"]


import io as _io
import time

import openpyxl as _openpyxl
from fastapi.testclient import TestClient
from app.main import app
from app import excel as _excel

_client = TestClient(app)
BORN = r"E:\PJ_OCR\test\3.資產負債表測試-此為可複製文字之PDF.pdf"


def test_excel_endpoint_builds_xlsx_from_sheets():
    r = _client.post("/api/excel", json={
        "file_name": "財報.pdf",
        "sheets": {"1": {"columns": ["項目"], "rows": [["現金"]]},
                   "2": {"columns": [], "rows": [["x"]]}},
    })
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    wb = _openpyxl.load_workbook(_io.BytesIO(r.content))
    assert wb.sheetnames == ["第1頁", "第2頁"]


def test_excel_endpoint_empty_sheets_400():
    r = _client.post("/api/excel", json={"file_name": "x.pdf", "sheets": {}})
    assert r.status_code == 400


def test_structure_endpoint_flow(monkeypatch):
    monkeypatch.setattr(_excel, "structure_page",
                        lambda text, key, **kw: {"columns": ["c"], "rows": [[text[:2]]]})
    with open(BORN, "rb") as f:
        jid = _client.post("/api/jobs",
                           files={"file": ("file3.pdf", f, "application/pdf")}).json()["job_id"]
    assert _client.post(f"/api/jobs/{jid}/structure").status_code == 409
    _client.post(f"/api/jobs/{jid}/recognize")
    for _ in range(60):
        if _client.get(f"/api/jobs/{jid}").json()["status"] in ("done", "error"):
            break
        time.sleep(0.3)
    assert _client.post(f"/api/jobs/{jid}/structure").status_code == 200
    s = None
    for _ in range(60):
        s = _client.get(f"/api/jobs/{jid}").json()
        if s["structure_status"] in ("done", "error"):
            break
        time.sleep(0.3)
    assert s["structure_status"] == "done"
    assert s["tables"]["1"]["columns"] == ["c"]
